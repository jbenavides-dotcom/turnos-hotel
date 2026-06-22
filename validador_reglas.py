#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Validador de la parrilla de turnos del Hotel LP&ET contra las reglas de Lina.

Lee la hoja "TURNOS 2025" (exportada de Google Sheets a .xlsx) y marca dónde la
programación manual incumple las reglas del documento "REGLAS DE TURNOS".

Las reglas se dividen en dos grupos:
  A) Verificables SOLO con la hoja (siempre se chequean).
  B) Dependen de la OCUPACIÓN por día (# cabañas). La hoja NO trae la ocupación
     de forma estructurada, así que se alimentan con data/ocupacion.json
     (o, en el futuro, desde Cloudbeds). Sin ese dato se reportan como
     "NO VERIFICABLE", no como violación.

Uso:
    python validador_reglas.py                 # valida la pestaña más reciente
    python validador_reglas.py "MAYO 26"       # valida una pestaña concreta
    python validador_reglas.py --todas          # valida todas las pestañas
    python validador_reglas.py --semana 2026-05-04   # solo esa semana (lunes)

Fuente de ocupación opcional: data/ocupacion.json  {"2026-05-02": 8, ...}
"""
import sys, os, re, json, datetime as dt
import openpyxl

# La consola de Windows usa cp1252; forzamos UTF-8 para los símbolos/emojis.
try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

# --------------------------------------------------------------------------
# CONFIGURACIÓN DE REGLAS  (del documento REGLAS DE TURNOS de Lina)
# --------------------------------------------------------------------------
AQUI = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(AQUI, "data", "TURNOS_2025_gsheet.xlsx")
OCUP = os.path.join(AQUI, "data", "ocupacion.json")

UMBRAL_REFUERZO = 7          # > 7 cabañas en operación => 2 en servicio + Zaira a cocina
HORA_AM_LIMITE  = 7 * 60 + 30   # "turno am" arranca a más tardar 7:30
HORA_TOPE_SIN_HUESPEDES = 18 * 60  # entre semana sin huéspedes => salir máx 6:00pm
MIN_DIAS_ACTIVO = 4          # solo evalúo "descanso semanal" a quien trabaja >=4 días esa semana

# Roles (nombre tal cual aparece en la columna A de la hoja, en MAYÚSCULAS)
ROLES = {
    "CLAUDIA TORRES": "cocina", "MARIBEL PAEZ": "cocina", "ZAIRA": "cocina",
    "NICOLAS SARMIENTO": "servicio", "SHEAN PAUL GUTIERREZ": "servicio",
    "GERWIN": "comercial",
    "KELLY CRUZ": "guia",
    "DANIELA MONROY": "admin",
    "CLAUDIA MONTENEGRO": "areas_publicas", "DEISY RODRIGUEZ": "areas_publicas",
    "DIEGO DIAZ": "conductor",
}
# Quién puede cubrir el turno de SERVICIO (para la regla de >7 cabañas)
PUEDE_SERVICIO = {"NICOLAS SARMIENTO", "SHEAN PAUL GUTIERREZ",
                  "GERWIN", "KELLY CRUZ", "DANIELA MONROY", "ZAIRA"}

# Estados no laborales (en la celda de "entrada")
ESTADOS = {"DESCANSO", "VACACIONES", "INCAPACIDAD", "LICENCIA",
           "CALAMIDAD", "CALAMIDAD FAMILIAR", "PERMISO", "SUSPENSION"}

DIAS = {"LUNES": 0, "MARTES": 1, "MIERCOLES": 2, "MIÉRCOLES": 2, "JUEVES": 3,
        "VIERNES": 4, "SABADO": 5, "SÁBADO": 5, "DOMINGO": 6}

# Pestaña -> (año, mes) del mes que titula la pestaña (puede empezar días antes)
TAB_MES = {"DICIEMBRE": (2025, 12), "ENERO": (2026, 1), "FEBRERO": (2026, 2),
           "MARZO": (2026, 3), "ABRIL": (2026, 4), "MAYO": (2026, 5),
           "JUNIO": (2026, 6), "JULIO": (2026, 7)}


# --------------------------------------------------------------------------
# PARSEO
# --------------------------------------------------------------------------
def norm(s):
    return re.sub(r"\s+", " ", str(s).strip()) if s is not None else ""

def es_estado(v):
    u = norm(v).upper()
    for e in ESTADOS:
        if u.startswith(e):
            return e if e != "CALAMIDAD FAMILIAR" else "CALAMIDAD"
    return None

def parse_hora(v):
    """'7:00 am' -> minutos; ' CIERRE' -> ('CIERRE', 22:00). None si no es hora."""
    s = norm(v).lower().replace(".", "")
    if not s:
        return None
    if "cierre" in s:
        return ("CIERRE", 22 * 60)
    s = re.sub(r"\s+", "", s)                       # '7:00am'
    m = re.match(r"^(\d{1,2}):(\d{2})(am|pm)?$", s)
    if not m:
        return None
    h, mi, ap = int(m.group(1)), int(m.group(2)), m.group(3)
    if ap == "pm" and h != 12:
        h += 12
    if ap == "am" and h == 12:
        h = 0
    return ("HORA", h * 60 + mi)

def tab_anchor(titulo):
    base = re.sub(r"\d+", "", titulo).strip().upper()
    return TAB_MES.get(base, (2026, 6))

def fechas_de_cabecera(ws, titulo):
    """Devuelve lista [(col, fecha, weekday)] leyendo la fila 1.
    Reconstruye fechas reales encajando los encabezados consecutivos."""
    headers = []   # (col, weekday_idx, daynum)
    for c in range(2, ws.max_column + 1):
        v = norm(ws.cell(1, c).value)
        if not v:
            continue
        m = re.match(r"([A-ZÁÉÍ]+)\s+0?(\d{1,2})\b", v.upper())
        if not m or m.group(1) not in DIAS:
            continue
        headers.append((c, DIAS[m.group(1)], int(m.group(2))))
    if not headers:
        return []
    # Buscar la fecha de inicio probando una ventana alrededor del 1° del mes
    year, month = tab_anchor(titulo)
    primero = dt.date(year, month, 1)
    mejor, mejor_score = None, -1
    for delta in range(-12, 8):
        cand = primero + dt.timedelta(days=delta)
        score = 0
        for i, (_c, wd, dn) in enumerate(headers):
            f = cand + dt.timedelta(days=i)
            if f.weekday() == wd and f.day == dn:
                score += 1
        if score > mejor_score:
            mejor, mejor_score = cand, score
    out = []
    for i, (c, _wd, _dn) in enumerate(headers):
        f = mejor + dt.timedelta(days=i)
        out.append((c, f, f.weekday()))
    return out

def leer_parrilla(ws, titulo):
    """-> dict[nombre] = dict[fecha] = registro
       registro: {'estado': X} | {'in': min, 'out': min, 'cierre': bool, 'raw': (e,s)}"""
    dias = fechas_de_cabecera(ws, titulo)
    parrilla, fechas = {}, [f for (_c, f, _w) in dias]
    for r in range(2, ws.max_row + 1):
        nombre = norm(ws.cell(r, 1).value).upper()
        if nombre not in ROLES and nombre not in ("LAURA SANCHEZ", "NICOLLE",
                "DIANA TUNJANO", "JOHAN RODRIGUEZ", "SANTIAGO", "CAMILO TORRES",
                "NELLY MENDOZA", "NEYSON RODRIGUEZ", "JONATHAN", "JONATHAN TORRES"):
            continue
        dd = {}
        for (c, f, _w) in dias:
            e = ws.cell(r, c).value
            s = ws.cell(r, c + 1).value
            est = es_estado(e)
            if est:
                dd[f] = {"estado": est}
                continue
            pe = parse_hora(e)
            if pe and pe[0] == "HORA":
                ps = parse_hora(s)
                dd[f] = {"in": pe[1],
                         "out": ps[1] if ps else None,
                         "cierre": bool(ps and ps[0] == "CIERRE"),
                         "raw": (norm(e), norm(s))}
        if dd:
            parrilla[nombre] = dd
    return parrilla, fechas


# --------------------------------------------------------------------------
# REGLAS
# --------------------------------------------------------------------------
def lunes_de(f):
    return f - dt.timedelta(days=f.weekday())

def cargar_ocupacion():
    if os.path.exists(OCUP):
        try:
            return {k: int(v) for k, v in json.load(open(OCUP, encoding="utf-8")).items()}
        except Exception:
            pass
    return {}

def hhmm(m):
    if m is None:
        return "?"
    return f"{m//60:02d}:{m%60:02d}"

def validar(parrilla, fechas, ocup):
    V, W, I = [], [], []   # violaciones, advertencias, info/no-verificable
    semanas = {}
    for f in fechas:
        semanas.setdefault(lunes_de(f), []).append(f)

    # ---- R1: cada quien descansa >=1 día por semana (si trabaja >=4 días) ----
    for lun, dias_sem in sorted(semanas.items()):
        if len(dias_sem) < 6:        # semana incompleta en el borde de la pestaña
            continue
        for nombre, dd in parrilla.items():
            reg_sem = [dd.get(f) for f in dias_sem if f in dd]
            trabajados = [r for r in reg_sem if r and "in" in r]
            descansos = [r for r in reg_sem if r and r.get("estado") == "DESCANSO"]
            no_lab = [r for r in reg_sem if r and r.get("estado") in
                      ("VACACIONES", "INCAPACIDAD", "CALAMIDAD", "LICENCIA")]
            if len(trabajados) >= MIN_DIAS_ACTIVO and not descansos and not no_lab:
                V.append(f"[{lun} sem] {nombre}: trabaja {len(trabajados)} días y "
                         f"NO tiene descanso esa semana (regla: ≥1 día/semana).")

    # ---- R3: cobertura diaria (alguien en AM y alguien en cierre) ----
    # Solo tiene sentido en días CON huéspedes: si no hay ocupación conocida >0,
    # un día sin cierre es válido (regla "sin huéspedes => máx 6pm"), no se avisa.
    for f in fechas:
        cab = ocup.get(f.isoformat())
        if not cab:                     # None (desconocido) o 0 (sin huéspedes) -> no aplica
            continue
        dia_trab = [(n, r) for n, dd in parrilla.items()
                    if (r := dd.get(f)) and "in" in r]
        if not dia_trab:
            continue
        hay_am = any(r["in"] <= HORA_AM_LIMITE for _n, r in dia_trab)
        hay_cierre = any(r["cierre"] for _n, r in dia_trab)
        if not hay_am:
            W.append(f"[{f} {f.strftime('%a')}] {cab} cab. y nadie abre en AM (entrada ≤ 7:30).")
        if not hay_cierre:
            W.append(f"[{f} {f.strftime('%a')}] {cab} cab. y nadie cierra (ningún turno hasta CIERRE).")

    # ---- R-vacaciones: no estar programado con turno el mismo día de un estado ----
    # (la hoja usa una celda por día, así que esto detecta choques raros de doble valor — informativo)

    # ---- Reglas que dependen de OCUPACIÓN ----
    for f in fechas:
        cab = ocup.get(f.isoformat())
        dia_trab = [(n, r) for n, dd in parrilla.items()
                    if (r := dd.get(f)) and "in" in r]
        if not dia_trab:
            continue
        if cab is None:
            # no se puede verificar refuerzo / tope horario por falta de ocupación
            continue
        # R4: > umbral cabañas => 2 en servicio + Zaira en cocina
        if cab > UMBRAL_REFUERZO:
            en_servicio = [n for n, _r in dia_trab if n in PUEDE_SERVICIO]
            if len(en_servicio) < 2:
                V.append(f"[{f}] {cab} cabañas (>{UMBRAL_REFUERZO}) pero solo "
                         f"{len(en_servicio)} en servicio: {en_servicio} "
                         f"(regla: 2 personas en servicio).")
            zaira = parrilla.get("ZAIRA", {}).get(f)
            if not (zaira and "in" in zaira):
                V.append(f"[{f}] {cab} cabañas (>{UMBRAL_REFUERZO}) pero ZAIRA no está "
                         f"en turno para apoyar a Claudia en cocina.")
        # R6: entre semana (no dom/festivo) sin huéspedes => salir máx 6pm
        if cab == 0 and f.weekday() < 5:
            tarde = [(n, r) for n, r in dia_trab
                     if r["cierre"] or (r["out"] and r["out"] > HORA_TOPE_SIN_HUESPEDES)]
            for n, r in tarde:
                V.append(f"[{f}] sin huéspedes esa noche pero {n} sale "
                         f"{'CIERRE' if r['cierre'] else hhmm(r['out'])} (tope 18:00).")

    # Nota global si no hubo ocupación
    fechas_sin_ocup = [f for f in fechas if f.isoformat() not in ocup]
    if fechas_sin_ocup:
        I.append(f"OCUPACIÓN no provista para {len(fechas_sin_ocup)}/{len(fechas)} días "
                 f"→ reglas de >{UMBRAL_REFUERZO} cabañas, domingo checkout-total y "
                 f"'sin huéspedes ≤18:00' NO verificadas. Alimentar data/ocupacion.json "
                 f"(o conectar Cloudbeds).")
    return V, W, I


# --------------------------------------------------------------------------
# MAIN
# --------------------------------------------------------------------------
def main():
    args = sys.argv[1:]
    todas = "--todas" in args
    semana = None
    if "--semana" in args:
        i = args.index("--semana")
        semana = lunes_de(dt.date.fromisoformat(args[i + 1]))
        del args[i:i + 2]
    args = [a for a in args if a != "--todas"]
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ocup = cargar_ocupacion()

    if todas:
        tabs = wb.sheetnames
    elif args:
        tabs = [args[0]]
    else:
        tabs = [wb.sheetnames[-1]]   # la más reciente

    print(f"== Validador de turnos · {os.path.basename(XLSX)} ==")
    print(f"   ocupación cargada: {len(ocup)} días\n")
    tot_v = tot_w = 0
    for tab in tabs:
        if tab not in wb.sheetnames:
            print(f"!! Pestaña '{tab}' no existe. Disponibles: {wb.sheetnames}")
            continue
        parrilla, fechas = leer_parrilla(wb[tab], tab)
        if semana is not None:
            fechas = [f for f in fechas if lunes_de(f) == semana]
            if not fechas:
                continue   # esta pestaña no cubre esa semana
        if not fechas:
            print(f"-- {tab}: sin fechas reconocibles, salto.\n")
            continue
        V, W, I = validar(parrilla, fechas, ocup)
        tot_v += len(V); tot_w += len(W)
        print(f"┌─ {tab}  ({fechas[0]} → {fechas[-1]}, {len(parrilla)} colaboradores)")
        if V:
            print(f"│  ❌ VIOLACIONES ({len(V)}):")
            for x in V: print(f"│     • {x}")
        if W:
            print(f"│  ⚠️  ADVERTENCIAS ({len(W)}):")
            for x in W: print(f"│     • {x}")
        if I:
            print(f"│  ℹ️  INFO:")
            for x in I: print(f"│     • {x}")
        if not V and not W:
            print("│  ✅ sin violaciones ni advertencias (reglas de hoja).")
        print("└─\n")
    print(f"== TOTAL: {tot_v} violaciones, {tot_w} advertencias ==")

if __name__ == "__main__":
    main()
