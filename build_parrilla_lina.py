#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Genera docs/parrilla_lina.json con la programación REAL de Lina (SOLO LECTURA).

Dos fuentes:
  - default: lee un .xlsx local (snapshot bajado a mano).
  - --sheet : lee la hoja en vivo vía Google Sheets API con la cuenta de servicio
              (para el GitHub Action). Credenciales: env GOOGLE_SA_KEY_JSON (JSON)
              o un key file local. Si la SA no tiene acceso (403), NO sobrescribe
              el snapshot existente y termina sin error.

NUNCA escribe en la hoja de Lina.

Salida: {"generated": "...", "plan": {"<empId>|<YYYY-MM-DD>": {tipo,ini,fin}}}
Uso:
  python build_parrilla_lina.py                  # xlsx local
  python build_parrilla_lina.py --sheet          # Sheets API (SA)
"""
import sys, os, json, datetime as dt
import openpyxl
import validador_reglas as V

AQUI = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(AQUI, "data", "TURNOS_2025_gsheet_LIVE.xlsx")
SALIDA = os.path.join(AQUI, "docs", "parrilla_lina.json")
SHEET_ID = "1ENHmx8JYq28URorBaX_AuVavB8r7fPT8tJa6C5UoEUE"
SA_FILES = [
    r"C:\Users\USUARIO\cafe-trazabilidad\service-account-key.json",
    r"C:\Users\USUARIO\.config\google-analytics\service-account-key.json",
]

NAME2ID = {
    "CLAUDIA TORRES": "ctorres", "MARIBEL PAEZ": "maribel", "ZAIRA": "zaira",
    "NICOLAS SARMIENTO": "nicolas", "SHEAN PAUL GUTIERREZ": "shean", "KELLY CRUZ": "kelly",
    "DANIELA MONROY": "daniela", "GERWIN": "gerwin", "CLAUDIA MONTENEGRO": "cmonte",
    "DEISY RODRIGUEZ": "deisy", "DIEGO DIAZ": "diego", "NICOLLE": "nicolle", "SANTIAGO": "santiago",
}


# --- adaptador para que las funciones de validador_reglas trabajen sobre una grilla de valores ---
class _Cell:
    __slots__ = ("value",)
    def __init__(self, v): self.value = v

class GridWS:
    def __init__(self, rows, title):
        self.rows, self.title = rows, title
        self.max_row = len(rows)
        self.max_column = max((len(r) for r in rows), default=0)
    def cell(self, r, c):
        try:
            return _Cell(self.rows[r - 1][c - 1])
        except IndexError:
            return _Cell(None)


def hhmm(m):
    return f"{m // 60:02d}:{m % 60:02d}"

def to_shift(e, s):
    est = V.es_estado(e)
    if est == "DESCANSO":
        return {"tipo": "descanso"}
    if est in ("VACACIONES", "CALAMIDAD", "INCAPACIDAD", "LICENCIA", "PERMISO", "SUSPENSION"):
        return {"tipo": "vacaciones"}
    pe = V.parse_hora(e)
    if not pe or pe[0] != "HORA":
        return None
    ini = hhmm(pe[1])
    ps = V.parse_hora(s)
    if ps and ps[0] == "CIERRE":
        return {"tipo": "tarde", "ini": ini, "fin": "cierre"}
    if ps and ps[0] == "HORA":
        return {"tipo": "manana", "ini": ini, "fin": hhmm(ps[1])}
    return {"tipo": "manana", "ini": ini, "fin": "15:00"}

def construir_plan(worksheets):
    """worksheets: lista de objetos con .cell(r,c).value, .max_row, .max_column, .title"""
    plan, sin_mapear = {}, set()
    for ws in worksheets:
        dias = V.fechas_de_cabecera(ws, ws.title)
        if not dias:
            continue
        for r in range(2, 30):
            nom = str(ws.cell(r, 1).value or "").strip().upper()
            if not nom:
                continue
            eid = NAME2ID.get(nom)
            for (c, f, _w) in dias:
                e = ws.cell(r, c).value
                if e is None or not str(e).strip():
                    continue
                sh = to_shift(e, ws.cell(r, c + 1).value)
                if sh is None:
                    continue
                if not eid:
                    sin_mapear.add(nom); continue
                plan[f"{eid}|{f.isoformat()}"] = sh
    return plan, sin_mapear


def fuente_xlsx():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    return [wb[t] for t in wb.sheetnames]  # ws de openpyxl ya tiene .cell/.max_*/.title

def fuente_sheets():
    """Lee la hoja en vivo con la SA. Devuelve lista de GridWS, o None si no hay acceso."""
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    info = os.environ.get("GOOGLE_SA_KEY_JSON")
    if info:
        creds = service_account.Credentials.from_service_account_info(
            json.loads(info), scopes=["https://www.googleapis.com/auth/spreadsheets.readonly"])
    else:
        kf = next((p for p in SA_FILES if os.path.exists(p)), None)
        if not kf:
            print("Sin credenciales de SA (GOOGLE_SA_KEY_JSON o key file)."); return None
        creds = service_account.Credentials.from_service_account_file(
            kf, scopes=["https://www.googleapis.com/auth/spreadsheets.readonly"])
    svc = build("sheets", "v4", credentials=creds, cache_discovery=False)
    try:
        meta = svc.spreadsheets().get(spreadsheetId=SHEET_ID, fields="sheets.properties.title").execute()
    except Exception as e:
        print("La SA no puede leer la hoja (¿falta compartir en modo Lector?):", str(e)[:140])
        return None
    out = []
    for sp in meta.get("sheets", []):
        title = sp["properties"]["title"]
        vals = svc.spreadsheets().values().get(
            spreadsheetId=SHEET_ID, range=f"'{title}'", valueRenderOption="FORMATTED_VALUE"
        ).execute().get("values", [])
        out.append(GridWS(vals, title))
    return out


def main():
    usar_sheets = "--sheet" in sys.argv
    if usar_sheets:
        worksheets = fuente_sheets()
        if worksheets is None:
            print("-> Conservo el snapshot existente (sin cambios)."); return
    else:
        worksheets = fuente_xlsx()

    plan, sin_mapear = construir_plan(worksheets)
    if not plan:
        print("-> Plan vacío; NO sobrescribo el snapshot."); return
    out = {"generated": dt.datetime.now().isoformat(timespec="seconds"), "plan": plan}
    os.makedirs(os.path.dirname(SALIDA), exist_ok=True)
    json.dump(out, open(SALIDA, "w", encoding="utf-8"), ensure_ascii=False, indent=0)
    print(f"-> {SALIDA}  ({len(plan)} celdas de Lina · fuente={'Sheets API' if usar_sheets else 'xlsx'})")
    if sin_mapear:
        print("  Sin id en la herramienta:", sorted(sin_mapear))

if __name__ == "__main__":
    main()
